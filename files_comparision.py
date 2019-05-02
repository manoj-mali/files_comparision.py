def excel_files(file1, file2):
    try:
        import openpyxl
        wb1 = openpyxl.load_workbook(file1)
        wb2 = openpyxl.load_workbook(file2)
        sheet1 = wb1.active
        sheet2 = wb2.active
        rows1 = sheet1.max_row
        rows2 = sheet2.max_row
        col1 = sheet1.max_column
        col2 = sheet2.max_column
        var = None
        for r1, r2 in zip(range(1, rows1 + 1), range(1, rows2 + 1)):
            for c1, c2 in zip(range(1, col1 + 1), range(1, col2 + 1)):
                if sheet1.cell(row=r1, column=c1).value == sheet2.cell(row=r2, column=c2).value:
                    print(end="")
                else:
                    print('diff in  ' + '(' + str(r1) + ', ' + str(c1) + ')' + ' of file1 and file2' + '\n', sheet1.cell(row=r1, column=c1).value, '\n', sheet2.cell(row=r2, column=c2).value)
                    var = 'Not none'
        if var == None:
            print("No diff")

    except ModuleNotFoundError:
        inp = input("You don't have 'openpyxl' module . Type 'y' to install ")
        if inp == 'y':
            import os
            try:
                os.system('pip3 install openpyxl')
            except:
                os.system('pip install openpyxl')
            print('----------------------------------------------------------------')
            import openpyxl

            wb1 = openpyxl.load_workbook(file1)
            wb2 = openpyxl.load_workbook(file2)
            sheet1 = wb1.active
            sheet2 = wb2.active
            rows1 = sheet1.max_row
            rows2 = sheet2.max_row
            col1 = sheet1.max_column
            col2 = sheet2.max_column
            var = None
            for r1, r2 in zip(range(1, rows1+1), range(1, rows2+1)):
                for c1, c2 in zip(range(1, col1+1), range(1, col2+1)):
                    if sheet1.cell(row=r1, column=c1).value == sheet2.cell(row=r2, column=c2).value:
                        print(end="")
                    elif sheet1.cell(row=r1, column=c1).value != sheet2.cell(row=r2, column=c2).value:
                        print('diff in  '+'('+str(r1)+', '+str(c1)+')'+' of file1 and file2'+'\n',sheet1.cell(row=r1, column=c1).value,'\n', sheet2.cell(row=r2, column=c2).value)
                        var = 'Not none'
            if var == None:
                print("No diff")

    except ImportError:
        inp = input("You don't have 'openpyxl' module . Type 'y' to install ")
        if inp == 'y':
            import os
            try:
                os.system('pip3 install openpyxl')
            except:
                os.system('pip install openpyxl')
            print('----------------------------------------------------------------')
            import openpyxl

            wb1 = openpyxl.load_workbook(file1)
            wb2 = openpyxl.load_workbook(file2)
            sheet1 = wb1.active
            sheet2 = wb2.active
            rows1 = sheet1.max_row
            rows2 = sheet2.max_row
            col1 = sheet1.max_column
            col2 = sheet2.max_column
            var = None
            for r1, r2 in zip(range(1, rows1+1), range(1, rows2+1)):
                for c1, c2 in zip(range(1, col1+1), range(1, col2+1)):
                    if sheet1.cell(row=r1, column=c1).value == sheet2.cell(row=r2, column=c2).value:
                        print(end="")
                    elif sheet1.cell(row=r1, column=c1).value != sheet2.cell(row=r2, column=c2).value:
                        print('diff in  '+'('+str(r1)+', '+str(c1)+')'+' of file1 and file2'+'\n',sheet1.cell(row=r1, column=c1).value,'\n', sheet2.cell(row=r2, column=c2).value)
                        var = 'Not none'
            if var == None:
                print("No diff")


def text_files(file1, file2):
    file1_m = open(file1, 'r')
    file2_m = open(file2, 'r')
    num_lines1 = sum(1 for count1 in open(file1))
    num_lines2 = sum(1 for count2 in open(file2))
    for line1 in range(1, num_lines1+1):
        for line2 in range(1, num_lines2 + 1):
            line_text1 = file1_m.readline()
            line_text2 = file2_m.readline()
            if line_text1 == line_text2:
                print(end="")
            elif line_text1 != line_text2:
                try:
                    with open(file1) as f1: data = f1.readlines()
                    line_no = data.index(line_text1)
                    print("diff in line num: "+str(line_no+1)+" of file 1")
                except:
                    with open(file2) as f2: data = f2.readlines()
                    line_no1 = data.index(line_text2)
                    print("diff in line num: "+str(line_no1+1)+" of file 2")
                print(line_text1 + line_text2)
